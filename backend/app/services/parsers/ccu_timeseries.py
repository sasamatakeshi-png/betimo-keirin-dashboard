"""同時接続数 時系列 Excel（YouTube監視ツール出力）のパーサ。

シート「データ」を読む。1行目ヘッダ:
    取得日時 / チャンネル名 / 番組タイトル / 同時接続数 / 登録者数 / URL / 種別
取得日時は JST のナイーブ datetime。URL の v= パラメータが動画ID。

シート「設定」には計測スケジュールが入る。C列ラベル「計測開始日時▶」の
右隣（D列）に計測開始日時（JST ナイーブ datetime）がある。これは「配信開始」
そのものではなく計測ウィンドウの開始だが、published_at が他に無いときの
フォールバック起点として使う（既存68点データもこの値を競合の起点にしている）。

返り値:
- parse_ccu_timeseries_xlsx(content) -> 時系列点のフラットなリスト（後方互換）
- parse_ccu_xlsx(content) -> {"start_time": datetime|None, "points": [...]}
各点: { "youtube_video_id", "channel_name", "title", "recorded_at"(JST aware),
        "value"(int), "subscriber_count"(int|None) }
同接が空・非数の行（取得エラー等）は含めない。
"""

from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone
from urllib.parse import parse_qs, urlparse

import openpyxl

JST = timezone(timedelta(hours=9))

_DATA_SHEET = "データ"
_SETTINGS_SHEET = "設定"
_START_LABEL = "計測開始日時"

# ヘッダ名 → 必須かどうか
_COLUMNS = {
    "recorded_at": ("取得日時", True),
    "channel_name": ("チャンネル名", True),
    "title": ("番組タイトル", False),
    "value": ("同時接続数", True),
    "subscriber_count": ("登録者数", False),
    "url": ("url", True),
}


def extract_video_id(url: str | None) -> str | None:
    """YouTube URL から動画IDを抽出（watch?v= / youtu.be/ の2形式に対応）。"""
    if not url:
        return None
    try:
        parsed = urlparse(str(url).strip())
    except ValueError:
        return None
    if parsed.netloc.endswith("youtu.be"):
        vid = parsed.path.lstrip("/").split("/")[0]
        return vid or None
    qs = parse_qs(parsed.query)
    vid = (qs.get("v") or [None])[0]
    return vid or None


def _read_data_points(ws) -> list[dict]:
    """「データ」シートのワークシートから同接時系列点を取り出す。"""
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower() if h is not None else "" for h in next(rows, [])]

    idx: dict[str, int] = {}
    for field, (name, required) in _COLUMNS.items():
        found = next((i for i, h in enumerate(headers) if name in h), None)
        if found is None and required:
            raise ValueError(f"ヘッダ「{name}」が見つかりません: {headers}")
        if found is not None:
            idx[field] = found

    def cell(row: tuple, field: str) -> object | None:
        i = idx.get(field)
        return row[i] if i is not None and i < len(row) else None

    points: list[dict] = []
    for row in rows:
        recorded = cell(row, "recorded_at")
        value = cell(row, "value")
        video_id = extract_video_id(cell(row, "url"))  # type: ignore[arg-type]
        if not isinstance(recorded, datetime) or video_id is None:
            continue
        if not isinstance(value, (int, float)):  # 取得エラー行などは同接が空
            continue
        sub = cell(row, "subscriber_count")
        points.append(
            {
                "youtube_video_id": video_id,
                "channel_name": str(cell(row, "channel_name") or "").strip(),
                "title": str(cell(row, "title") or "").strip(),
                # 取得日時はJSTのナイーブ datetime → JST aware に変換
                "recorded_at": recorded.replace(tzinfo=JST),
                "value": int(round(float(value))),
                "subscriber_count": int(round(float(sub)))
                if isinstance(sub, (int, float))
                else None,
            }
        )
    return points


def _read_measurement_start(wb) -> datetime | None:
    """「設定」シートから計測開始日時（JST aware）を読む。無ければ None。"""
    if _SETTINGS_SHEET not in wb.sheetnames:
        return None
    ws = wb[_SETTINGS_SHEET]
    for row in ws.iter_rows(values_only=True):
        label_idx = next(
            (i for i, c in enumerate(row) if isinstance(c, str) and _START_LABEL in c),
            None,
        )
        if label_idx is None:
            continue
        # ラベルの右側で最初に現れる datetime を計測開始日時とみなす
        for c in row[label_idx + 1 :]:
            if isinstance(c, datetime):
                return c.replace(tzinfo=JST)
    return None


def parse_ccu_timeseries_xlsx(content: bytes) -> list[dict]:
    """「データ」シートから同接時系列点を取り出す（後方互換のフラットリスト）。"""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    if _DATA_SHEET not in wb.sheetnames:
        raise ValueError(f"シート「{_DATA_SHEET}」が見つかりません: {wb.sheetnames}")
    points = _read_data_points(wb[_DATA_SHEET])
    wb.close()
    return points


def parse_ccu_xlsx(content: bytes) -> dict:
    """計測開始日時と時系列点をまとめて返す。

    返り値: {"start_time": datetime|None(JST aware), "points": list[dict]}
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    if _DATA_SHEET not in wb.sheetnames:
        raise ValueError(f"シート「{_DATA_SHEET}」が見つかりません: {wb.sheetnames}")
    start_time = _read_measurement_start(wb)
    points = _read_data_points(wb[_DATA_SHEET])
    wb.close()
    return {"start_time": start_time, "points": points}
